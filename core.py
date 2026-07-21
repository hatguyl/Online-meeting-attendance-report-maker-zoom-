import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from rapidfuzz import fuzz
import re
import os

def extract_roll_name(text):
    if not text:
        return "", ""

    text = str(text).strip()

    cleaned = re.sub(r'[^A-Za-z0-9 ]+', ' ', text)

    match = re.search(r'\d+[A-Za-z]+\d+', cleaned)

    if match:
        roll = match.group().upper()
        name = cleaned.replace(match.group(), "").strip()
        return roll, name

    return "", cleaned.strip()


def is_red_or_yellow(cell):
    fill = cell.fill
    if not fill or not fill.start_color:
        return False

    rgb = fill.start_color.rgb
    if not rgb:
        return False

    rgb = rgb.upper()

    RED = "00FF7F7F"
    YELLOW = "00FFD966"
    GREEN = "0090EE90"

    if rgb == GREEN:
        return False

    return rgb in (RED, YELLOW)


def extract_att_data(file_path):
    wb = load_workbook(file_path)
    ws = wb["ATT"]

    data = []

    for row in ws.iter_rows(min_row=2):
        name_cell = row[0]
        duration_cell = row[1]

        if not name_cell.value:
            continue

        if is_red_or_yellow(duration_cell):
            roll, name = extract_roll_name(name_cell.value)

            data.append({
                "Roll Number": roll,
                "Name": name,
                "Issue Type": "LESS_TIME"
            })

    return data


def extract_late_data(file_path):
    df = pd.read_excel(file_path, sheet_name="LATE")

    data = []

    for value in df.iloc[:, 0]:
        if pd.isna(value):
            continue

        roll, name = extract_roll_name(value)

        data.append({
            "Roll Number": roll,
            "Name": name,
            "Issue Type": "LATE_ARRIVAL"
        })

    return data


def export_to_csv(data, save_path, file_name):
    if not data:
        return None

    df = pd.DataFrame(data)

    df = df.rename(columns={
        "Roll Number": "roll_number",
        "Name": "student_name",
        "Issue Type": "issue_type"
    })

    df = df[["roll_number", "student_name", "issue_type"]]
    df.fillna("", inplace=True)

    full_path = os.path.join(save_path, f"{file_name}.csv")
    df.to_csv(full_path, index=False)

    return full_path

# +----------------+
# | EXCEL STYLES   |
# +----------------+
HEADER_FILL = PatternFill("solid", fgColor="FFE699")
META_FILL = PatternFill("solid",fgColor="FFC000")
HEADER_FONT = Font(color="000000", bold=True)

ROW1 = PatternFill("solid",fgColor="FFFFFF")
ROW2 = PatternFill("solid",fgColor="FFF2CC")

GREEN  = PatternFill("solid", fgColor="90EE90")
RED    = PatternFill("solid", fgColor="FF7F7F")
YELLOW = PatternFill("solid", fgColor="FFD966")

# +----------------+
# | REGEX / CONST  |
# +----------------+
ROLL_PATTERN = re.compile(
    r'\b(\d{1,2}[A-Z]{1,4}\s*\d{2,4})\b'   # 10CB168, 10CA041
    r'|'
    r'\b([A-Z]{2,4}\s*\d{3,4})\b',           # CA026 (min 3 digits avoids "FF10")
    re.IGNORECASE
)
NON_WORD_RE = re.compile(r"[^\w\s]", re.UNICODE)
PAREN_RE = re.compile(r"\(.*?\)")
APOSTROPHE_RE = re.compile(r"[’‘`]")

BLACKLIST_WORDS = {"mentorbee", "mentor", "classroom", "dont", "don't"}
TITLE_WORDS = {"miss", "sir"}
UNKNOWN_WORDS = {
    "zoom", "user", "iphone", "ipad", "realme",
    "redmi", "oppo", "vivo", "samsung", "galaxy","xiaomi",
    "desktop", "laptop", "windows", "android",
    "phone", "tab", "device", "guest", "participant",
    "lenovo", "mi", "moto", "oneplus"
}

# +----------------+
# | HELPERS        |
# +----------------+
def find_column(df, keys):

    aliases = {
        "join": [
            "join",
            "join time",
            "joined",
            "participant joined",
            "attended from"
        ],

        "leave": [
            "leave",
            "leave time",
            "left",
            "participant left",
            "attended to"
        ],

        "duration": [
            "duration",
            "time in meeting",
            "minutes",
            "attendance duration"
        ],

        "name": [
            "name",
            "participant",
            "participant name",
            "display name",
            "user"
        ]
    }

    expanded = []

    for k in keys:
        expanded.extend(aliases.get(k, [k]))

    for c in df.columns:

        lc = c.lower().strip()

        if any(a in lc for a in expanded):
            return c

    raise ValueError(f"Missing required column: {keys}")

def parse_datetime(col):
    return pd.to_datetime(col, errors="coerce", format="mixed")

def extract_meeting_metadata(csv_file):

    try:

        raw = pd.read_csv(
            csv_file,
            header=None,
            nrows=5
        ).fillna("")

        meta = {
            "topic": "",
            "meeting_id": "",
            "duration": "",
            "start": "",
            "end": "",
            "participants": ""
        }

        row1 = raw.iloc[0].tolist()
        row2 = raw.iloc[1].tolist()

        for i, key in enumerate(row1):

            k = str(key).strip().lower()

            value = ""
            if i < len(row2):
                value = row2[i]

            if "topic" in k:
                meta["topic"] = value

            elif k == "id":
                meta["meeting_id"] = value

            elif "duration" in k:
                meta["duration"] = value

            elif "start" in k:
                meta["start"] = value

            elif "end" in k:
                meta["end"] = value

            elif "participant" in k:
                meta["participants"] = value

        return meta

    except Exception as e:

        print("Metadata extraction failed:", e)

        return {
            "topic": "",
            "meeting_id": "",
            "duration": "",
            "start": "",
            "end": "",
            "participants": ""
        }
    
def group_csv_files(csv_files):

    groups = {}

    for path in csv_files:

        meta = extract_meeting_metadata(path)

        topic = meta["topic"]

        start = meta["start"]

        try:
            start = pd.to_datetime(start, format="mixed")
            rounded = start.strftime("%Y-%m-%d %H")
        except:
            rounded = "UNKNOWN"

        key = (topic, rounded)

        groups.setdefault(key, []).append(path)

    return list(groups.values())

def combine_csv_files(csv_files):

    if isinstance(csv_files, str):
        csv_files = [csv_files]

    dfs = []

    for path in csv_files:

        loaded = False

        # try multiple header rows
        for header_row in range(0, 15):

            try:
                df = pd.read_csv(
                    path,
                    header=header_row
                )

                df.columns = [
                    str(c).strip()
                    for c in df.columns
                ]

                try:
                    find_column(df, {"join"})
                    find_column(df, {"leave"})
                    find_column(df, {"name"})

                    dfs.append(df)

                    loaded = True
                    print(
                        f"Loaded CSV: {path} "
                        f"(header row {header_row})"
                    )

                    break

                except:
                    continue

            except Exception:
                continue

        if not loaded:
            print("Skipped invalid CSV:", path)

    if not dfs:
        raise ValueError(
            "No valid Zoom participant CSV files found"
        )

    return pd.concat(dfs, ignore_index=True)
    
def format_datetime_12h(dt):
    if pd.isna(dt):
        return ""
    return dt.strftime("%m/%d/%Y %I:%M:%S %p")

# Regex that normalises separator characters (-, _, .) sitting between
# a roll-number token and a name token so that ROLL_PATTERN's \b fires.
# e.g. "10CB028-Jerome" -> "10CB028 Jerome"
#      "10CB106.Tanvi"  -> "10CB106 Tanvi"
#      "10cb017_ANTONIO"-> "10cb017 ANTONIO"
_SEP_AFTER_DIGITS  = re.compile(r'(?<=\d)([-_.])(?=[A-Za-z])')
_SEP_BEFORE_DIGITS = re.compile(r'(?<=[A-Za-z])([-_.])(?=\d)')
# Catches rolls glued directly to name letters with no separator at all.
# e.g. "10CA058Aakarsh", "Niranjan10CB163", "10CB126bryan"
_BARE_ROLL_RE = re.compile(r'(\d{1,2}[A-Za-z]{1,4}\d{2,4})', re.IGNORECASE)

def _sep_normalise(s: str) -> str:
    """Replace roll↔name separator punctuation with a plain space."""
    s = _SEP_AFTER_DIGITS.sub(' ', s)
    s = _SEP_BEFORE_DIGITS.sub(' ', s)
    return s

def extract_roll(name):
    if not name:
        return None
    s = str(name).strip()

    # 1. Roll embedded in parentheses: "Aibel Shanto (10CA041)", "(10CA060)Avelino"
    for paren_content in re.findall(r'\(([^)]+)\)', s):
        m = _BARE_ROLL_RE.search(paren_content)
        if m:
            return m.group(1).upper().replace(' ', '')

    # 2. Normalise separators so word-boundary pattern works
    s_norm = _sep_normalise(s)

    # 3. Standard word-boundary ROLL_PATTERN on normalised string
    for m in ROLL_PATTERN.finditer(s_norm):
        val = m.group(1) or m.group(2)
        return val.upper().replace(' ', '').replace('_', '').replace('-', '')

    # 4. Last resort: any digit-letter-digit sequence anywhere in the string
    #    (handles no-space, name-before-roll, separator-less variants)
    m = _BARE_ROLL_RE.search(s_norm)
    if m:
        return m.group(1).upper().replace(' ', '')

    return None

def normalize_name(name):
    if not name:
        return ""
    s = str(name).strip()

    # Remove parenthetical groups (aliases, device names, duplicate roll in brackets)
    s = PAREN_RE.sub("", s)

    # Normalise roll↔name separators so ROLL_PATTERN can strip the roll cleanly
    s = _sep_normalise(s)

    # Strip roll number via word-boundary pattern
    stripped = ROLL_PATTERN.sub("", s)

    # If the roll was glued (no-space / no separator), strip it with the bare regex
    if re.search(r'\d{2,}', stripped):
        stripped = _BARE_ROLL_RE.sub("", stripped)

    # Remove remaining non-word punctuation and collapse whitespace
    stripped = NON_WORD_RE.sub("", stripped)
    stripped = re.sub(r"\s+", " ", stripped)
    return stripped.strip().lower()

def fuzzy_match(a, b, threshold=82):
    if not a or not b:
        return False

    score = fuzz.ratio(a.lower(), b.lower())
    return score >= threshold

def extract_alias(name):
    if not name:
        return None
    m = re.search(r"\(([^)]+)\)", str(name))
    return normalize_name(m.group(1)) if m else None

def is_blacklisted(name):
    if not name:
        return False
    s = APOSTROPHE_RE.sub("'", str(name).lower())
    if any(w in s for w in BLACKLIST_WORDS):
        return True
    parts = s.split()
    return len(parts) >= 2 and parts[-1] in TITLE_WORDS

def is_unknown_identity(name):

    if extract_roll(name):
        return False
    
    if not name:
        return False

    s = normalize_name(name)

    score = 0

    words = set(s.split())

    # ---------- DEVICE / GENERIC WORDS ----------
    matches = words & UNKNOWN_WORDS
    score += len(matches) * 35
    device_patterns = [
        ("realme", "pad"),
        ("galaxy", "tab"),
        ("redmi", "note"),
        ("iphone",),
        ("ipad",),
    ]

    for pattern in device_patterns:
        if all(p in words for p in pattern):
            score += 40

    # ---------- VERY SHORT ----------
    if len(s) <= 3:
        score += 40

    # ---------- NO SPACES ----------
    if " " not in s:
        score += 15

    if re.search(r"[a-z]+\d+", s):
        score += 20

    generic = {
        "zoom user",
        "guest user",
        "unknown user",
        "participant",
        "student"
    }

    if s in generic:
        score += 60

    return score >= 50

def format_table(ws):
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        col[0].fill = HEADER_FILL
        col[0].font = HEADER_FONT
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = width + 3
    for r in range(2, ws.max_row + 1):
        fill = ROW1 if r % 2 == 0 else ROW2
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = fill

# +----------------+
# | MAIN PROCESS   |
# +----------------+
def process(
    csv_files,
    do_att,
    do_late,
    late_time,
    late_ampm,
    total_att,
    absent_limit,
    out_path,
    *,
    selected_data="",
    batch_name="",
    issued_date=""
):
    # ---------- LOAD ----------
    if isinstance(csv_files, str):
        csv_files = [csv_files]

    df = combine_csv_files(csv_files)
    df.columns = df.columns.str.strip()
    
    name_col  = find_column(df, {"name", "participant", "user"})
    join_col  = find_column(df, {"join"})
    leave_col = find_column(df, {"leave"})
    dur_col   = find_column(df, {"duration"})

    df["_JOIN_DT"]  = parse_datetime(df[join_col])
    df["_LEAVE_DT"] = parse_datetime(df[leave_col])
    df["_DUR"] = pd.to_numeric(df[dur_col], errors="coerce").fillna(0)

    df["_ROLL"]      = df[name_col].apply(extract_roll)
    df["_NAME_KEY"]  = df[name_col].apply(normalize_name)
    df["_ALIAS_KEY"] = df[name_col].apply(extract_alias)

    df_original = df.copy()

    # ---------- CF / BLACKLIST ----------
    cf_mask = df_original[name_col].apply(is_blacklisted)
    cf = df_original.loc[cf_mask].copy()

    # ---------- UNKNOWN ----------
    unknown_mask = df_original[name_col].apply(is_unknown_identity)
    unknown = df_original.loc[
        unknown_mask & ~cf_mask
    ].copy()

    # ---------- CLEAN DATA ----------
    df = df_original.loc[
        ~(cf_mask | unknown_mask)
    ].copy()
    # ---------- MEETING DURATION ----------
    vt = df.dropna(subset=["_JOIN_DT", "_LEAVE_DT"])
    meeting_minutes = (
        vt["_LEAVE_DT"].max() - vt["_JOIN_DT"].min()
    ).total_seconds() / 60.0 if not vt.empty else 0

    # ---------- PERSON MODEL ----------
    class Person:

        def improve_display(self, other_display):
            if len(str(other_display)) > len(str(self.display)):
                self.display = other_display

        def __init__(self, r):
            self.roll = r["_ROLL"]
            self.name = r["_NAME_KEY"]
            self.alias = r["_ALIAS_KEY"]
            self.display = r[name_col]
            self.duration = r["_DUR"]
            self.first_join = r["_JOIN_DT"]

        def priority_name(self):
            # brackets > roll > plain
            if "(" in self.display and ")" in self.display:
                return (0, self.display)
            if self.roll:
                return (1, self.display)
            return (2, self.display)

    persons = []

    # ---------- LEVEL 1: EXACT / NEAR-EXACT DUPLICATES ----------
    # Merges rows that are clearly the same session entry:
    # same roll + same name + same join time (exact), OR
    # same roll + same name with no join time gap > 5 min (handles Zoom's 2-sec drift).
    for _, r in df.iterrows():
        matched = False
        for p in persons:
            same_identity = (p.roll and p.roll == r["_ROLL"] and p.name == r["_NAME_KEY"]) or \
                            (not p.roll and not r["_ROLL"] and p.name == r["_NAME_KEY"] and p.first_join == r["_JOIN_DT"])
            if same_identity:
                try:
                    gap = abs((p.first_join - r["_JOIN_DT"]).total_seconds())
                except Exception:
                    gap = 9999
                if gap < 300:  # within 5 minutes = same session slot
                    p.duration += r["_DUR"]
                    if pd.notna(r["_JOIN_DT"]) and (pd.isna(p.first_join) or r["_JOIN_DT"] < p.first_join):
                        p.first_join = r["_JOIN_DT"]
                    matched = True
                    break
        if not matched:
            persons.append(Person(r))

    # ---------- LEVEL 1.5: NO-ROLL NAME DUPLICATES ----------
    # Merges bare-name rows into existing persons (rolled or not) by fuzzy name match.
    # Rolled rows go first so no-roll rows (e.g. "Aibel Shanto") find their rolled match.
    merged = []
    rolled_first = [p for p in persons if p.roll] + [p for p in persons if not p.roll]
    for p in rolled_first:
        if not p.roll:
            found = False
            for m in merged:
                if fuzzy_match(m.name, p.name):
                    m.duration += p.duration
                    try:
                        if pd.notna(p.first_join) and (pd.isna(m.first_join) or p.first_join < m.first_join):
                            m.first_join = p.first_join
                    except Exception:
                        pass
                    found = True
                    break
            if not found:
                merged.append(p)
        else:
            merged.append(p)

    persons = merged


    # ---------- LEVEL 2: BRACKET ALIAS ----------
    # Merges rows where the bracketed alias text clearly refers to another entry.
    # Threshold 90 to avoid near-misses like "vismaya" → "vismaiy" (different people).
    # No join-time gate — alias match is explicit intent, not coincidence.
    merged = []

    for p in persons:
        done = False

        for m in merged:

            alias_match = False

            # alias -> name
            if p.alias and fuzzy_match(p.alias, m.name, 90):
                alias_match = True

            # name -> alias
            elif m.alias and fuzzy_match(p.name, m.alias, 90):
                alias_match = True

            # alias -> alias
            elif p.alias and m.alias and fuzzy_match(p.alias, m.alias, 90):
                alias_match = True

            if alias_match:
                m.duration += p.duration

                # ---------- KEEP BEST ROLL ----------
                if not m.roll and p.roll:
                    m.roll = p.roll

                # ---------- KEEP BEST DISPLAY ----------
                if p.roll:
                    m.display = p.display
                else:
                    m.improve_display(p.display)

                # ---------- KEEP EARLIEST JOIN ----------
                try:
                    if pd.notna(p.first_join) and (pd.isna(m.first_join) or p.first_join < m.first_join):
                        m.first_join = p.first_join
                except Exception:
                    pass

                done = True
                break

        if not done:
            merged.append(p)

    persons = merged

    # ---------- LEVEL 3: ROLL BASED ----------
    # Same roll + similar name → same person. Name check at 60 prevents false merges
    # when two different students share a roll due to data entry errors.
    final = []
    for p in persons:
        merged_flag = False
        for f in final:
            if p.roll and f.roll and p.roll == f.roll:
                if fuzzy_match(p.name, f.name, 60):
                    f.duration += p.duration
                    if p.roll:
                        f.display = p.display
                    else:
                        f.improve_display(p.display)
                    try:
                        if pd.notna(p.first_join) and (pd.isna(f.first_join) or p.first_join < f.first_join):
                            f.first_join = p.first_join
                    except Exception:
                        pass
                    merged_flag = True
                    break

        if not merged_flag:
            final.append(p)

    # ---------- FINAL TABLE ----------
    rows = []
    for p in final:
        rows.append({
            "Name (original name)": p.display,
            "Duration (minutes)": round(p.duration, 2),
            "First Join": p.first_join,
            "_PRIORITY": p.priority_name()
        })

    out = pd.DataFrame(rows)

    out = out.sort_values(
        by=["_PRIORITY", "Name (original name)"],
        na_position="last"
    )
    out.drop(columns="_PRIORITY", inplace=True)


    absent = out[out["Duration (minutes)"] < absent_limit]

    # ---------- LATE ----------
    cutoff = datetime.datetime.strptime(
            f"{late_time} {late_ampm}",
            "%I:%M:%S %p"
        ).time()
    late = out[out["First Join"].dt.time > cutoff].copy()
    late = late.sort_values(
        by=["Name (original name)", "First Join"],
        ascending=[True, True],
        na_position="last"
    )
    late["Join Time"] = late["First Join"].apply(format_datetime_12h)
    late["Status"] = "LATE"
    late = late[["Name (original name)", "Join Time", "Status"]]

    # ---------- RAW ----------
    raw = df_original[
        [name_col, join_col, leave_col, dur_col]
    ].copy()

    raw["Join Time"] = df_original["_JOIN_DT"].apply(
        format_datetime_12h
    )

    raw.rename(
        columns={name_col: "Name (original name)"},
        inplace=True
    )

    raw["_JOIN_SORT"] = df_original["_JOIN_DT"]

    raw = raw.sort_values(
        by=["Name (original name)", "_JOIN_SORT"],
        ascending=[True, True],
        na_position="last"
    )

    raw = raw[
        ["Name (original name)", "Join Time", leave_col, dur_col]
    ]

    host_words = [
        "mentorbee",
        "host",
        "co-host"
    ]

    raw = raw[
        ~raw["Name (original name)"]
        .astype(str)
        .str.lower()
        .str.contains("|".join(host_words), na=False)
    ]

    # ---------- CF ----------
    cf["Join Time"] = cf["_JOIN_DT"].apply(format_datetime_12h)
    cf = cf[[name_col, "Join Time", leave_col, dur_col]]
    cf.rename(columns={name_col: "Name (original name)"}, inplace=True)

    # ---------- UNKNOWN ----------
    unknown["Join Time"] = unknown["_JOIN_DT"].apply(format_datetime_12h)

    unknown = unknown[
        [name_col, "Join Time", leave_col, dur_col]
    ]

    unknown.rename(
        columns={name_col: "Name (original name)"},
        inplace=True
    )

    # ---------- WRITE ----------

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        raw.to_excel(
            writer,
            sheet_name="RAW",
            index=False,
            header=False,
            startrow=5
        )
        ws = writer.sheets["RAW"]

        from openpyxl.styles import Alignment

        # ==========================================
        # REPORT METADATA
        # ==========================================

        meta = extract_meeting_metadata(csv_files[0])

        metadata = [
            (
                [
                    "Topic",
                    "ID",
                    "Duration (minutes)",
                    "Start time",
                    "End time",
                    "Participants",
                    "Issued date"
                ],
                [
                    meta.get("topic", ""),
                    meta.get("meeting_id", ""),
                    meta.get("duration", ""),
                    str(meta.get("start", "")),
                    str(meta.get("end", "")),
                    len(raw),
                    issued_date
                ]
            ),
            (
                [
                    "Selected data",
                    "Given batch timing",
                    "Given attendance threshold",
                    "Given absent threshold",
                    "Given late cutoff"
                ],
                [
                    selected_data,
                    batch_name,
                    total_att,
                    absent_limit,
                    f"{late_time} {late_ampm}"
                ]
            )
        ]

        row = 1

        for headers, values in metadata:

            for col, text in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col, value=text)
                c.fill = META_FILL
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="left")

            row += 1

            for col, text in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=text)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="left")

            row += 1

        # ---------- TABLE HEADERS ----------
        headers = [
            "Name (original name)",
            "Join time",
            "Leave time",
            "Duration (minutes)"
        ]

        for col, value in enumerate(headers, start=1):
            ws.cell(row=5, column=col, value=value)

        if do_att:
            out[["Name (original name)", "Duration (minutes)"]].to_excel(
                writer, sheet_name="ATT", index=False
            )
            if not absent.empty:
                absent[["Name (original name)", "Duration (minutes)"]].to_excel(
                    writer, sheet_name="ABSENT", index=False
                )

        if do_late:
            late.to_excel(writer, sheet_name="LATE", index=False)

        if not cf.empty:
            cf.to_excel(writer, sheet_name="CF", index=False)
        if not unknown.empty:
            unknown.to_excel(
                writer,
                sheet_name="UNKNOWN",
                index=False
            )

    # ---------- FORMAT & HIGHLIGHT ----------
    wb = load_workbook(out_path)
    for s in wb.sheetnames:

        if s == "RAW":
            continue

        format_table(wb[s])

    # ---------- RAW METADATA STYLE ----------
    ws = wb["RAW"]

    # ---------- PARTICIPANT HEADERS ----------
    ws.auto_filter.ref = f"A5:D{ws.max_row}"
    for cell in ws[5]:
        cell.fill = HEADER_FILL
        cell.font = Font(
            color="000000",
            bold=True
        )

    # ==========================================
    # RAW COLUMN WIDTHS
    # ==========================================

    for col in ws.columns:

        max_len = 0

        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[
            col[0].column_letter
        ].width = min(max_len + 2, 45)

    # ==========================================
    # RAW ROW STRIPING
    # ==========================================

    for r in range(7, ws.max_row + 1):

        for c in range(1, ws.max_column + 1):

            cell = ws.cell(r, c)

            if cell.fill == GREEN or cell.fill == RED or cell.fill == YELLOW:
                continue

            fill = ROW1 if r % 2 == 0 else ROW2
            cell.fill = fill

    # ==========================================
    # FREEZE
    # ==========================================

    ws.freeze_panes = "A6"

    # RAW highlights (respect user selections)
    ws = wb["RAW"]
    for r in range(6, ws.max_row + 1):
        
        # ---- LATE highlight (Join Time) ----
        if do_late:
            try:
                t = datetime.datetime.strptime(
                    ws[f"B{r}"].value,
                    "%m/%d/%Y %I:%M:%S %p"
                ).time()
                ws[f"B{r}"].fill = GREEN if t <= cutoff else RED
            except:
                pass

        # ---- ATT / ABSENT highlight (Duration) ----
        if do_att:
            try:
                d = float(ws[f"D{r}"].value)
                ws[f"D{r}"].fill = (
                    YELLOW if d < absent_limit else
                    RED if d < total_att else
                    GREEN
                )
            except:
                pass

    # ATT / ABSENT highlights
    if do_att:
        ws = wb["ATT"]
        for r in range(2, ws.max_row + 1):
            d = float(ws[f"B{r}"].value)
            ws[f"B{r}"].fill = (
                YELLOW if d < absent_limit else
                RED if d < total_att else
                GREEN
            )
        if "ABSENT" in wb.sheetnames:
            ws = wb["ABSENT"]
            for r in range(2, ws.max_row + 1):
                ws[f"B{r}"].fill = YELLOW

    # LATE highlights
    if do_late and "LATE" in wb.sheetnames:
        ws = wb["LATE"]
        for r in range(2, ws.max_row + 1):
            ws[f"B{r}"].fill = RED
            ws[f"C{r}"].fill = RED

    wb.save(out_path)
